"""
Excel Generator
Creates minimalist, clean Excel files with configurable styling
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any

class ExcelGenerator:
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.styling = config.get("excel_styling", {})
        self.custom_colors = config.get("colors", {}).get("custom_colors", {})
        
    def create_timetable(self, schedule_df: pd.DataFrame, output_path: str) -> bool:
        """Create main timetable Excel file"""
        try:
            if schedule_df.empty:
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Timetable"
            
            # Create timetable grid
            self._create_timetable_grid(ws, schedule_df)
            
            # Apply styling
            self._apply_styling(ws, schedule_df)
            
            # Save file
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return True
            
        except Exception:
            return False
    
    def create_summary_workbook(self, schedule_df: pd.DataFrame, output_path: str) -> bool:
        """Create workbook with multiple summary sheets"""
        try:
            if schedule_df.empty:
                return False
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create different view sheets
            self._create_instructor_sheet(wb, schedule_df)
            self._create_room_sheet(wb, schedule_df)
            self._create_section_sheet(wb, schedule_df)
            self._create_program_sheet(wb, schedule_df)
            self._create_overview_sheet(wb, schedule_df)
            
            # Save file
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return True
            
        except Exception:
            return False
    
    def _create_timetable_grid(self, ws, schedule_df: pd.DataFrame):
        """Create the main timetable grid"""
        # Get unique time slots and sort them
        time_slots = self._get_time_slots(schedule_df)
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        rooms = sorted(schedule_df["Room"].unique())
        
        # Headers
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(time_slots))
        ws["A1"] = "University Timetable"
        
        # Column headers
        headers = ["Day", "Room"] + time_slots
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
        
        # Populate data
        current_row = 3
        
        for day in days:
            day_schedule = schedule_df[schedule_df["Weekday"] == day]
            if day_schedule.empty:
                continue
            
            day_rooms = sorted(day_schedule["Room"].unique())
            day_start_row = current_row
            
            for room in day_rooms:
                # Day and room columns
                ws.cell(row=current_row, column=1, value=day)
                ws.cell(row=current_row, column=2, value=room)
                
                # Time slot data
                room_schedule = day_schedule[day_schedule["Room"] == room]
                
                for col, time_slot in enumerate(time_slots, start=3):
                    cell = ws.cell(row=current_row, column=col)
                    session = self._find_session_for_slot(room_schedule, time_slot)
                    
                    if session is not None:
                        cell_text = self._format_session_text(session)
                        cell.value = cell_text
                
                current_row += 1
            
            # Merge day cells if multiple rooms
            if current_row > day_start_row + 1:
                ws.merge_cells(start_row=day_start_row, start_column=1, 
                              end_row=current_row - 1, end_column=1)
    
    def _create_instructor_sheet(self, wb, schedule_df: pd.DataFrame):
        """Create instructor summary sheet"""
        ws = wb.create_sheet("Instructors")
        
        headers = ["Instructor", "Course", "Section", "Program", "Day", "Time", "Room"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        row = 2
        for instructor in sorted(schedule_df["Instructor"].unique()):
            instructor_schedule = schedule_df[schedule_df["Instructor"] == instructor]
            
            for _, session in instructor_schedule.sort_values(["Weekday", "Start_Time"]).iterrows():
                ws.cell(row=row, column=1, value=instructor)
                ws.cell(row=row, column=2, value=session["Course_Name"])
                ws.cell(row=row, column=3, value=session["Section"])
                ws.cell(row=row, column=4, value=session["Program"])
                ws.cell(row=row, column=5, value=session["Weekday"])
                ws.cell(row=row, column=6, value=f"{session['Start_Time']}-{session['End_Time']}")
                ws.cell(row=row, column=7, value=session["Room"])
                row += 1
        
        self._style_summary_sheet(ws)
    
    def _create_room_sheet(self, wb, schedule_df: pd.DataFrame):
        """Create room summary sheet"""
        ws = wb.create_sheet("Rooms")
        
        headers = ["Room", "Day", "Time", "Course", "Instructor", "Section", "Program"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        row = 2
        for room in sorted(schedule_df["Room"].unique()):
            room_schedule = schedule_df[schedule_df["Room"] == room]
            
            for _, session in room_schedule.sort_values(["Weekday", "Start_Time"]).iterrows():
                ws.cell(row=row, column=1, value=room)
                ws.cell(row=row, column=2, value=session["Weekday"])
                ws.cell(row=row, column=3, value=f"{session['Start_Time']}-{session['End_Time']}")
                ws.cell(row=row, column=4, value=session["Course_Name"])
                ws.cell(row=row, column=5, value=session["Instructor"])
                ws.cell(row=row, column=6, value=session["Section"])
                ws.cell(row=row, column=7, value=session["Program"])
                row += 1
        
        self._style_summary_sheet(ws)
    
    def _create_section_sheet(self, wb, schedule_df: pd.DataFrame):
        """Create section summary sheet"""
        ws = wb.create_sheet("Sections")
        
        headers = ["Section", "Program", "Course", "Instructor", "Day", "Time", "Room"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        row = 2
        for section in sorted(schedule_df["Section"].unique()):
            section_schedule = schedule_df[schedule_df["Section"] == section]
            
            for _, session in section_schedule.sort_values(["Program", "Weekday", "Start_Time"]).iterrows():
                ws.cell(row=row, column=1, value=section)
                ws.cell(row=row, column=2, value=session["Program"])
                ws.cell(row=row, column=3, value=session["Course_Name"])
                ws.cell(row=row, column=4, value=session["Instructor"])
                ws.cell(row=row, column=5, value=session["Weekday"])
                ws.cell(row=row, column=6, value=f"{session['Start_Time']}-{session['End_Time']}")
                ws.cell(row=row, column=7, value=session["Room"])
                row += 1
        
        self._style_summary_sheet(ws)
    
    def _create_program_sheet(self, wb, schedule_df: pd.DataFrame):
        """Create program summary sheet"""
        ws = wb.create_sheet("Programs")
        
        headers = ["Program", "Section", "Course", "Instructor", "Day", "Time", "Room"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        row = 2
        for program in sorted(schedule_df["Program"].unique()):
            program_schedule = schedule_df[schedule_df["Program"] == program]
            
            for _, session in program_schedule.sort_values(["Section", "Weekday", "Start_Time"]).iterrows():
                ws.cell(row=row, column=1, value=program)
                ws.cell(row=row, column=2, value=session["Section"])
                ws.cell(row=row, column=3, value=session["Course_Name"])
                ws.cell(row=row, column=4, value=session["Instructor"])
                ws.cell(row=row, column=5, value=session["Weekday"])
                ws.cell(row=row, column=6, value=f"{session['Start_Time']}-{session['End_Time']}")
                ws.cell(row=row, column=7, value=session["Room"])
                row += 1
        
        self._style_summary_sheet(ws)
    
    def _create_overview_sheet(self, wb, schedule_df: pd.DataFrame):
        """Create overview statistics sheet"""
        ws = wb.create_sheet("Overview")
        
        # Statistics
        stats = [
            ["Total Sessions", len(schedule_df)],
            ["Lab Sessions", len(schedule_df[schedule_df["Is_Lab"]])],
            ["Theory Sessions", len(schedule_df[~schedule_df["Is_Lab"]])],
            ["Instructors", schedule_df["Instructor"].nunique()],
            ["Rooms Used", schedule_df["Room"].nunique()],
            ["Programs", schedule_df["Program"].nunique()],
            ["Sections", schedule_df["Section"].nunique()]
        ]
        
        ws.cell(row=1, column=1, value="Schedule Statistics")
        for row, (label, value) in enumerate(stats, start=3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
        
        # Daily distribution
        ws.cell(row=1, column=4, value="Daily Distribution")
        daily_dist = schedule_df.groupby("Weekday").size()
        for row, (day, count) in enumerate(daily_dist.items(), start=3):
            ws.cell(row=row, column=4, value=day)
            ws.cell(row=row, column=5, value=count)
        
        self._style_summary_sheet(ws)
    
    def _apply_styling(self, ws, schedule_df: pd.DataFrame):
        """Apply minimalist styling to worksheet"""
        # Colors
        header_color = self.styling.get("header_color", "4472C4")
        day_color = self.styling.get("day_color", "D9E2F3")
        room_color = self.styling.get("room_color", "F2F2F2")
        
        # Fonts
        font_name = self.styling.get("font_name", "Calibri")
        font_size = self.styling.get("font_size", 10)
        
        # Create styles
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        day_fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
        room_fill = PatternFill(start_color=room_color, end_color=room_color, fill_type="solid")
        
        header_font = Font(name=font_name, size=font_size + 2, bold=True, color="FFFFFF")
        regular_font = Font(name=font_name, size=font_size)
        
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        
        # Apply to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = regular_font
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        
        # Style headers
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Style title
        ws["A1"].font = Font(name=font_name, size=font_size + 4, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # Style day and room columns
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = day_fill  # Day column
            ws.cell(row=row, column=2).fill = room_fill  # Room column
        
        # Color-code sessions
        for row in range(3, ws.max_row + 1):
            for col in range(3, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    course_name = self._extract_course_name(cell.value)
                    if course_name:
                        color = self._get_course_color(course_name)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Auto-adjust columns if enabled
        if self.styling.get("auto_adjust_columns", True):
            self._adjust_column_widths(ws)
    
    def _style_summary_sheet(self, ws):
        """Apply styling to summary sheets"""
        font_name = self.styling.get("font_name", "Calibri")
        font_size = self.styling.get("font_size", 10)
        header_color = self.styling.get("header_color", "4472C4")
        
        # Style headers
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(name=font_name, size=font_size, bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust columns
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _get_time_slots(self, schedule_df: pd.DataFrame) -> list:
        """Extract and sort unique time slots"""
        time_slots = set()
        for _, row in schedule_df.iterrows():
            slot = f"{row['Start_Time']} - {row['End_Time']}"
            time_slots.add(slot)
        
        return sorted(list(time_slots), key=lambda x: x.split(" - ")[0])
    
    def _find_session_for_slot(self, room_schedule: pd.DataFrame, time_slot: str):
        """Find session that matches the time slot"""
        slot_start, slot_end = time_slot.split(" - ")
        
        for _, session in room_schedule.iterrows():
            if session["Start_Time"] <= slot_start < session["End_Time"]:
                return session
        
        return None
    
    def _format_session_text(self, session) -> str:
        """Format session information for display"""
        lab_text = " [LAB]" if session["Is_Lab"] else ""
        return f"{session['Course_Name']}\n{session['Section']}\n{session['Instructor']}{lab_text}"
    
    def _extract_course_name(self, cell_value: str) -> str:
        """Extract course name from cell value"""
        if not cell_value:
            return ""
        
        lines = str(cell_value).split("\n")
        return lines[0] if lines else ""
    
    def _get_course_color(self, course_name: str) -> str:
        """Get color for course (custom or generated)"""
        # Check for custom color first
        if course_name in self.custom_colors:
            return self.custom_colors[course_name]
        
        # Generate simple color based on hash
        if self.config.get("colors", {}).get("auto_generate", True):
            hash_val = hash(course_name) % 16777215  # 24-bit color space
            return f"{hash_val:06X}"
        
        return "FFFFFF"  # Default white
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths"""
        # Day column
        ws.column_dimensions["A"].width = 12
        # Room column
        ws.column_dimensions["B"].width = 15
        
        # Time slot columns
        for col in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Set row heights
        for row in range(1, ws.max_row + 1):
            if row == 1:
                ws.row_dimensions[row].height = 30  # Title
            elif row == 2:
                ws.row_dimensions[row].height = 25  # Headers
            else:
                ws.row_dimensions[row].height = 60  # Data rows