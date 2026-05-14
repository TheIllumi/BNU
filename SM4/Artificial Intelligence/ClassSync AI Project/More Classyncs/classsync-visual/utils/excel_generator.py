"""
Excel Generator for ClassSync Visual
Creates formatted Excel timetables from grid data
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any
import traceback

class ExcelGenerator:
    """Generates formatted Excel timetables"""
    
    def __init__(self, config_manager):
        self.config_manager = config_manager
        self.weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    def create_timetable(self, 
                        grid: Dict[str, Any], 
                        time_slots: List[str], 
                        output_path: Path,
                        title: str) -> Dict[str, Any]:
        """Create Excel timetable from grid data"""
        
        result = {'success': False, 'error': None}
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Timetable"
            
            # Setup styling
            self.setup_styles()
            
            # Create timetable
            self.create_header(ws, title, len(time_slots))
            self.create_column_headers(ws, time_slots)
            self.populate_timetable_data(ws, grid, time_slots)
            self.add_watermark(ws, len(time_slots))
            self.adjust_formatting(ws, time_slots)
            
            # Save workbook
            wb.save(output_path)
            result['success'] = True
            
        except Exception as e:
            result['error'] = str(e)
            print(f"Error creating Excel file: {traceback.format_exc()}")
        
        return result
    
    def setup_styles(self):
        """Setup reusable styles"""
        styling = self.config_manager.get('styling', {})
        
        # Border style with fallback
        border_style = styling.get('border_style', 'thin')
        self.border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style)
        )
        
        # Header styles with fallbacks
        header_color = styling.get('header_color', 'FF4472C4')
        self.header_fill = PatternFill(
            start_color=header_color,
            end_color=header_color,
            fill_type="solid"
        )
        
        font_name = styling.get('font_name', 'Calibri')
        header_font_size = styling.get('header_font_size', 12)
        self.header_font = Font(
            name=font_name,
            size=header_font_size,
            bold=True,
            color="FFFFFF"
        )
        
        # Day and room styles with fallbacks
        day_color = styling.get('day_color', 'FFD9E2F3')
        self.day_fill = PatternFill(
            start_color=day_color,
            end_color=day_color,
            fill_type="solid"
        )
        
        room_color = styling.get('room_color', 'FFF2F2F2')
        self.room_fill = PatternFill(
            start_color=room_color,
            end_color=room_color,
            fill_type="solid"
        )
        
        # Content fonts with fallbacks
        font_size = styling.get('font_size', 10)
        self.content_font = Font(
            name=font_name,
            size=font_size
        )
        
        self.bold_font = Font(
            name=font_name,
            size=font_size,
            bold=True
        )
    
    def create_header(self, ws, title: str, num_time_slots: int):
        """Create main header"""
        # Merge cells for header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + num_time_slots)
        
        # Set header text
        header_text = f"ClassSync Visual - {title}"
        ws['A1'] = header_text
        
        # Get styling with fallbacks
        styling = self.config_manager.get('styling', {})
        font_name = styling.get('font_name', 'Calibri')
        header_font_size = styling.get('header_font_size', 12)
        
        ws['A1'].font = Font(
            name=font_name,
            size=header_font_size + 2,
            bold=True
        )
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    def create_column_headers(self, ws, time_slots: List[str]):
        """Create column headers"""
        headers = ["Day", "Room"] + time_slots
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def populate_timetable_data(self, ws, grid: Dict[str, Any], time_slots: List[str]):
        """Populate timetable with data"""
        current_row = 3
        export_options = self.config_manager.get('export_options')
        
        for day in self.weekdays:
            day_rooms = grid.get(day, {})
            
            # Skip empty days if configured
            if not day_rooms and not export_options['include_empty_days']:
                continue
            
            # If no rooms for this day, add empty row
            if not day_rooms:
                day_rooms = {"": {slot: None for slot in time_slots}}
            
            day_start_row = current_row
            
            for room, slots in day_rooms.items():
                # Day column
                day_cell = ws.cell(row=current_row, column=1, value=day)
                
                # Room column
                room_cell = ws.cell(row=current_row, column=2, value=room)
                room_cell.fill = self.room_fill
                room_cell.font = self.bold_font
                room_cell.alignment = Alignment(horizontal='center', vertical='center')
                room_cell.border = self.border
                
                # Time slot columns
                for col, slot in enumerate(time_slots, start=3):
                    cell = ws.cell(row=current_row, column=col)
                    session = slots.get(slot)
                    
                    if session:
                        # Format session text
                        cell_text = self.format_session_text(session, export_options)
                        cell.value = cell_text
                        
                        # Apply course color
                        course_color = self.config_manager.get_color_for_course(session['course'])
                        cell.fill = PatternFill(
                            start_color=course_color,
                            end_color=course_color,
                            fill_type="solid"
                        )
                        
                        # Text formatting
                        cell.font = self.content_font
                        cell.alignment = Alignment(
                            wrap_text=True,
                            vertical='center',
                            horizontal='center'
                        )
                    else:
                        cell.value = ""
                    
                    cell.border = self.border
                
                current_row += 1
            
            # Merge day cells if enabled and there are multiple rooms
            if (export_options['merge_day_cells'] and 
                current_row > day_start_row + 1):
                
                ws.merge_cells(
                    start_row=day_start_row,
                    start_column=1,
                    end_row=current_row - 1,
                    end_column=1
                )
                
                # Style merged day cell
                day_cell = ws.cell(row=day_start_row, column=1)
                day_cell.fill = self.day_fill
                
                # Get styling with fallbacks
                styling = self.config_manager.get('styling', {})
                font_name = styling.get('font_name', 'Calibri')
                font_size = styling.get('font_size', 10)
                
                day_cell.font = Font(
                    name=font_name,
                    size=font_size + 1,
                    bold=True
                )
                day_cell.alignment = Alignment(horizontal='center', vertical='center')
                day_cell.border = self.border
            
            # Style individual day cells if not merged
            else:
                for row_idx in range(day_start_row, current_row):
                    day_cell = ws.cell(row=row_idx, column=1)
                    day_cell.fill = self.day_fill
                    day_cell.font = self.bold_font
                    day_cell.alignment = Alignment(horizontal='center', vertical='center')
                    day_cell.border = self.border
    
    def format_session_text(self, session: Dict[str, Any], export_options: Dict[str, Any]) -> str:
        """Format session text based on options"""
        if export_options['compact_mode']:
            # Compact format
            text_parts = [session['course']]
            if session['section']:
                text_parts.append(session['section'])
        else:
            # Full format
            text_parts = [session['course']]
            
            if session['section']:
                text_parts.append(session['section'])
            
            if export_options['show_program'] and session['program']:
                text_parts.append(session['program'])
            
            if export_options['show_instructor'] and session['instructor']:
                text_parts.append(session['instructor'])
        
        return '\n'.join(text_parts)
    
    def add_watermark(self, ws, num_time_slots: int):
        """Add watermark to worksheet"""
        watermark_config = self.config_manager.get('watermark')
        
        if not watermark_config['show']:
            return
        
        # Ensure color is in correct format
        watermark_color = "808080"  # Gray color for watermark
        
        watermark_font = Font(
            name="Arial",
            size=10,
            italic=True,
            color=watermark_color
        )
        
        if watermark_config['position'] == "top-header":
            # Update header text with watermark
            ws['A1'] = watermark_config['text']
            ws['A1'].font = watermark_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        else:  # bottom-right
            # Find last row with data
            last_row = ws.max_row + 2
            last_col = get_column_letter(2 + num_time_slots)
            
            watermark_cell = f"{last_col}{last_row}"
            ws[watermark_cell] = watermark_config['text']
            ws[watermark_cell].font = watermark_font
            ws[watermark_cell].alignment = Alignment(horizontal='right', vertical='center')
    
    def adjust_formatting(self, ws, time_slots: List[str]):
        """Adjust column widths and row heights"""
        styling = self.config_manager.get('styling', {})
        
        auto_adjust = styling.get('auto_adjust_columns', True)
        if not auto_adjust:
            return
        
        # Day column
        ws.column_dimensions['A'].width = 12
        
        # Room column
        ws.column_dimensions['B'].width = 15
        
        # Time slot columns
        for i, slot in enumerate(time_slots, start=3):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 25
        
        # Set row heights
        for row in range(1, ws.max_row + 1):
            if row == 1:  # Header row
                ws.row_dimensions[row].height = 30
            elif row == 2:  # Column headers
                ws.row_dimensions[row].height = 40
            else:  # Data rows
                ws.row_dimensions[row].height = 60